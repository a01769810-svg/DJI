r"""
sysid_gimbal.py — IDENTIFICACION DEL GIMBAL (tilt de camara) del DJI Neo, EN TIERRA.

POR QUE ES DISTINTO DEL YAW (leer antes de tocar nada):
  1. TODO EN TIERRA, SIN VOLAR. Inclinar la camara NO toca los motores (0x04/0x01 va al
     modulo gimbal, no al FC de vuelo). No hay --fly aqui: esta herramienta NO PUEDE
     despegar. Sin bateria en juego, sin riesgo -> se puede repetir gratis.
  2. EL GIMBAL TIENE TOPES MECANICOS (-90 .. +60 en el Neo del usuario, y NUESTRO camino
     los alcanza: la FASE A llego a -89.9 y a +56 antes de que cortara la guarda => misma
     autoridad que la app). El yaw giraba libre y un PRBS podia deambular sin consecuencias.
     Aqui, si la excursion supera el recorrido, el PRBS se queda pegado al limite y NO mide
     nada. (Tocar un tope NO es peligroso: la FASE 0 conduce contra ellos a proposito, que
     es lo mismo que hace la app al mantener la rueda de tilt. El coste es de DATOS.) Por eso:
       - FASE 0 mide los topes ANTES de nada (empujando suave y parando en cuanto el
         angulo deja de responder).
       - La amplitud del PRBS se calcula CON LOS DATOS REALES de la FASE A, no a ojo.

LA PLANTA (MEDIDA, EXP-037). Se esperaba "igual que el yaw"; NO lo es:
  0x04/0x01 es un comando de VELOCIDAD (uint16 en 363..1685; 1024=quieta, <1024=abajo,
  >1024=arriba). El feedback 0x04/0x05 da el ANGULO (gpitch, int16/10 => 0.1 deg).
  Hay integrador, si -- pero ademas:
    - ZONA MUERTA (50) Y EXPO (n=1.704) A LA VEZ. Por eso fallan los modelos simples: la
      potencia pura da R2=0.967 y la lineal+zona muerta R2=0.963; cada uno captura media
      verdad. Juntos: R2=0.999998. (El yaw era expo PURO, sin zona muerta.)
    - TAU=0.200 s OBSERVABLE (con tau R2=0.919, sin tau R2=0.322). En el yaw la diferencia
      era 0.00006 -> alli tau no existia a efectos practicos.
  => planta MAS DIFICIL que el yaw: polo extra en 0.6065 y ganancia local que varia x5.8
     (el yaw solo x1.9).
  => la I sigue sobrando (hay integrador; en el yaw costo 12.5% de overshoot y un ciclo
     limite, EXP-036), la D aqui SI aporta (hay polo que amortiguar), y la INVERSION de
     zona muerta+expo pasa de opcional (yaw) a RECOMENDADA -> hoja 'datos_lin'.

MUESTREO: el Push Position 0x04/0x05 llega a 9.90 Hz (medido sobre 'Novena captura.pcap':
679 paquetes / 68.5 s) => Ts = 100 ms, el MISMO que el OSD, y por el mismo motivo no se
puede subir: es la tasa a la que el dron reporta.

EXCURSION DEL PRBS: la suma acumulada de una m-secuencia de 7 bits esta ACOTADA en +-7
bits (no hace random walk). Luego la excursion es predecible:
      excursion = rate(amplitud) * Tb * 7
y se dimensiona para que quepa holgada en el recorrido medido en la FASE 0.

SECUENCIA
  FASE 0  topes: empuja suave abajo hasta que gpitch deja de moverse, luego arriba. Da el
          RECORRIDO real y el centro. Para en cuanto no responde (no insiste contra el tope).
  FASE A  escalones: para cada nivel |u| aplica +u y luego -u (simetrico => vuelve solo al
          punto de partida) y mide la velocidad por regresion sobre el 60% CENTRAL de cada
          escalon -> CURVA ESTATICA. (Un barrido de tope a tope seria inviable: a 26 deg/s
          de fondo, u=100 da 0.37 deg/s y cruzar 150 deg tardaria mas de 5 minutos.)
  FASE B  PRBS bang-bang centrado, con la amplitud CALCULADA de la FASE A para que la
          excursion quepa en el recorrido. -> DINAMICA / retardo.

USO
  .\neo.ps1 sysid_gimbal.py --probe     SOLO la FASE 0 (topes). La primera vez, correr esto.
  .\neo.ps1 sysid_gimbal.py             completo: topes + barridos + PRBS -> sysid_gimbal.csv
  .\.venv\Scripts\python tools\neo_control\sysid_gimbal.py --export sysid_gimbal.csv
Ctrl+C = para y centra la camara.
"""
import argparse, csv, os, sys, time, socket

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import neo_udp as N
import flight as F
from sysid_yaw import prbs7

TS = 0.1                 # impuesto por el Push Position 0x04/0x05 (9.90 Hz medido)
CENTER = 1024            # 0x04/0x01: 1024 = quieta
MAX_CMD = 660            # rango 363..1685 => +-660 sobre el centro (mismo que el stick)
PRBS_EXC = 7             # excursion acotada de una m-secuencia de 7 bits, en bits

# --- PLANTA DEL GIMBAL, MEDIDA EN TIERRA (EXP-037). NO es la misma forma que el yaw. ---
# ESTATICA: tiene zona muerta Y expo a la vez. Por eso fallaban los modelos simples
# (potencia pura R2=0.967, lineal+zona muerta R2=0.963); los dos capturaban media verdad:
#     rate = 0.000471454 * sign(u) * (|u| - 50)^1.704  deg/s   (0 si |u| < 50)
#     R2 = 0.999998, err_max 0.02 deg/s sobre los 6 escalones. Fondo de escala 26.2 deg/s.
# DINAMICA: rate[k] = 0.6065*rate[k-1] + 0.3935*r_sp[k-2] ; gpitch[k] = gpitch[k-1]+Ts*rate[k]
#     tau = 0.200 s y ES OBSERVABLE (con tau: R2=0.919; con tau->0: R2=0.322). En el yaw
#     esa diferencia era 0.00006 -> alli tau no existia a efectos practicos.
# => el gimbal es una planta MAS DIFICIL que el yaw: polo extra en 0.6065 y ganancia local
#    que varia x5.8 (u=100 -> 0.013 ; u=660 -> 0.073). El yaw solo variaba x1.9.
# => la I sigue sobrando (hay integrador), pero la D aqui SI aporta (hay polo que amortiguar)
#    y la INVERSION de zona muerta+expo pasa de opcional (yaw) a RECOMENDADA.
DEADZONE_FIT = 50.0
EXPO_A_FIT = 0.000471454
EXPO_N_FIT = 1.704
GIMBAL_TAU = 0.200
GIMBAL_DELAY = 1        # muestras, UNA VEZ modelado tau (sin tau, un retardo puro finge 2)


def rate_of(u):
    """Velocidad de la camara para un comando u. Curva MEDIDA (R2=0.999998)."""
    s = 1.0 if u >= 0 else -1.0
    return s * EXPO_A_FIT * max(abs(u) - DEADZONE_FIT, 0.0) ** EXPO_N_FIT


def cmd_for_rate(r):
    """INVERSION: comando u que produce la velocidad 'r' deg/s. Linealiza la planta."""
    if abs(r) < 1e-9:
        return 0.0
    s = 1.0 if r >= 0 else -1.0
    return s * (DEADZONE_FIT + (abs(r) / EXPO_A_FIT) ** (1.0 / EXPO_N_FIT))


def clamp(v, lo, hi):
    return max(lo, min(hi, v))


class Gimbal:
    """Envia velocidad al gimbal y lee su angulo, manteniendo viva la sesion.
    NO toca los motores de vuelo: 0x04/0x01 va al modulo gimbal."""

    def __init__(self, f):
        self.f = f
        self.nt = {"cmd": 0.0, "sub": 0.0, "ka": 0.0}
        self.gp = None            # ultimo gpitch leido
        self.t_gp = None

    def pump(self, u):
        """Manda 'u' (deflexion sobre 1024) a ~10Hz y drena el downlink.
        Devuelve (gpitch, t) si llego una muestra NUEVA, o (None, None)."""
        now = time.time()
        if self.f.serial and now >= self.nt["sub"]:
            self.f.sub13(); self.nt["sub"] = now + 0.2
        if now >= self.nt["ka"]:
            self.f.s.keepalive(); self.nt["ka"] = now + 0.5
        if now >= self.nt["cmd"]:
            self.f.gimbal_rate(CENTER + int(clamp(u, -MAX_CMD, MAX_CMD)))
            self.nt["cmd"] = now + TS
        self.f.s.sock.settimeout(0.02)
        try:
            d, a = self.f.s.sock.recvfrom(65535)
        except (socket.timeout, BlockingIOError):
            return None, None
        if not (d and a[0] == N.DRONE[0]):
            return None, None
        g = N.find_gimbal_position(d)
        if not g:
            return None, None
        self.gp, self.t_gp = g["gpitch"], time.time()
        return self.gp, self.t_gp

    def hold(self, secs, u=0):
        t0 = time.time()
        while time.time() - t0 < secs:
            self.pump(u)
        return self.gp

    def wait_reading(self, secs=4.0):
        t0 = time.time()
        while time.time() - t0 < secs:
            g, _ = self.pump(0)
            if g is not None:
                return g
        return None


# ------------------------------------------------------------------ FASE 0: topes
def find_limit(gb, u, args, samples, label):
    """Empuja a velocidad 'u' hasta que el angulo DEJA DE MOVERSE => tope mecanico.
    Para en cuanto no responde: NO insiste contra el tope. Devuelve el angulo del tope."""
    t0 = time.time()
    last_move = time.time()
    last = gb.gp
    while time.time() - t0 < args.limit_timeout:
        g, t = gb.pump(u)
        if g is None:
            continue
        samples.append(dict(t_s=round(time.time() - args.t0, 4), u_cmd=int(u),
                            phase="LIMIT_" + label, gpitch=g))
        if last is None or abs(g - last) > args.still_deg:
            last_move = time.time(); last = g
        elif time.time() - last_move > args.still_secs:
            print("   tope %s en %.1f deg (dejo de moverse %.1fs)" % (label, g, args.still_secs),
                  flush=True)
            gb.hold(0.3, 0)                       # suelta: no seguir empujando contra el tope
            return g
    print("   !! %s: no encontro tope en %.0fs (angulo %.1f) — se usa este valor"
          % (label, args.limit_timeout, gb.gp), flush=True)
    gb.hold(0.3, 0)
    return gb.gp


def goto(gb, target, args, samples, tol=1.5, timeout=None):
    """Lleva la camara a 'target' con un P sencillo (el esquema de point_camera).
    Solo para posicionar entre fases; no se identifica con esto.

    OJO: UN SOLO pump() por vuelta. pump() solo transmite en su ranura de 10Hz, asi que
    llamarlo dos veces (una para leer, otra para mandar) es una CARRERA: gana la primera
    y si esa mandaba 0 la camara no se mueve. Ese bug dejaba el PRBS sin centrar y lo
    estampaba contra el tope."""
    # el timeout se ESCALA a la distancia: un gimbal lento (~4.5 deg/s a fondo segun el
    # log de mapflight) tarda 33s en cruzar 150 deg. Un timeout fijo de 10s lo dejaba a
    # medias y descentraba el PRBS.
    if timeout is None:
        d0 = abs(target - gb.gp) if gb.gp is not None else 90.0
        timeout = clamp(d0 / 1.5 + 5.0, 8.0, 90.0)
    t0 = time.time()
    while time.time() - t0 < timeout:
        g = gb.gp                                  # ultima lectura conocida (no consume ranura)
        if g is None:
            gb.pump(0)
            continue
        err = target - g
        if abs(err) <= tol:
            gb.hold(0.2, 0)
            return g
        u = clamp(abs(err) * 40, 130, 400)
        gnew, t = gb.pump(u if err > 0 else -u)    # <- el UNICO pump de la vuelta
        if gnew is not None:
            samples.append(dict(t_s=round(t - args.t0, 4), u_cmd=0,
                                phase="GOTO", gpitch=gnew))
    print("   !! goto(%.1f) no llego: se quedo en %.1f" % (target, gb.gp or 0), flush=True)
    return gb.gp


# ------------------------------------------------------------------ FASE A: barridos
def step(gb, u, lo, hi, args, samples):
    """Aplica 'u' durante --step-secs y mide la velocidad estacionaria por regresion sobre
    el 60% CENTRAL (fuera transitorios de arranque y parada). Es la curva estatica.

    Escalon de DURACION FIJA cerca del centro, NO barrido de tope a tope: un barrido
    completo es inviable si el eje es lento (a 4.5 deg/s cruzar 150 deg son 33 s, y a
    u=100 mas de 5 min). Con escalones simetricos (+u y luego -u) se vuelve solo al punto
    de partida y sirve igual si el gimbal resulta rapido.
    Guarda: si nos acercamos a un tope, corta y devuelve lo medido hasta ahi."""
    pts = []
    t_end = time.time() + args.step_secs
    while time.time() < t_end:
        g = gb.gp
        if g is not None and ((u > 0 and g > hi - args.margin) or (u < 0 and g < lo + args.margin)):
            print("   (guarda: cerca del tope en %.1f, corto el escalon u=%+d)" % (g, u), flush=True)
            break
        gnew, t = gb.pump(u)                      # UN solo pump por vuelta
        if gnew is None:
            continue
        pts.append((t, gnew))
        samples.append(dict(t_s=round(t - args.t0, 4), u_cmd=int(u),
                            phase="STEP%+d" % u, gpitch=gnew))
    gb.hold(0.3, 0)
    if len(pts) < 8:
        return None
    a, b = int(len(pts) * 0.2), int(len(pts) * 0.8)
    seg = pts[a:b]
    if len(seg) < 4:
        return None
    ts = [p[0] - seg[0][0] for p in seg]
    gs = [p[1] for p in seg]
    n = len(seg)
    mt, mg = sum(ts) / n, sum(gs) / n
    den = sum((x - mt) ** 2 for x in ts)
    if den <= 0:
        return None
    return sum((ts[i] - mt) * (gs[i] - mg) for i in range(n)) / den   # deg/s


# ------------------------------------------------------------------ main
def run(args):
    args.t0 = time.time()
    samples = []
    print("=" * 68)
    print("  sysid_gimbal.py — IDENTIFICACION DEL GIMBAL (EN TIERRA, sin volar)")
    print("  Ts = 100ms (Push Position 0x04/0x05 a 9.90 Hz)  |  cmd 363..1685, centro 1024")
    print("  Esta herramienta NO despega: 0x04/0x01 va al gimbal, no al FC de vuelo.")
    print("=" * 68)

    s = N.Type5Session()
    if not s.open():
        print("!! sin ACK al hello -> revisa WiFi del Neo / DJI Fly cerrado."); return
    print("hello -> ACK. Sesion abierta.")
    f = F.Flight(s, None, None)
    gb = Gimbal(f)
    try:
        print("enviando init (%d frames)..." % len(F.INIT))
        for fr in F.INIT:
            f.s.send_command(fr); time.sleep(0.03)
        ok = f.engage()
        print(">>> ENGANCHE 0x51: %s" % ("OK" if ok else "FALLO (sin serial)"), flush=True)

        g0 = gb.wait_reading()
        if g0 is None:
            print("!! no llega el Push Position 0x04/0x05. Abortado."); return
        print("angulo inicial: %.1f deg" % g0, flush=True)

        # ---------------- FASE 0: topes ----------------
        if args.lo is not None and args.hi is not None:
            # topes DECLARADOS por el usuario (los suyos: -90/+60, comprobados con la app).
            # OJO: lo que permite la APP no tiene por que ser lo que alcanza NUESTRO camino
            # (DJI suele gatear el tilt hacia arriba tras un ajuste). Por eso --probe sigue
            # siendo lo recomendado la primera vez: confirma que tenemos la misma autoridad.
            lo, hi = args.lo, args.hi
            print("\n1) FASE 0 SALTADA: topes declarados %.1f .. %.1f  (--lo/--hi)" % (lo, hi))
            print("   OJO: sin verificar que nuestro path los alcance. Si el gimbal no llega,")
            print("   la excursion calculada no cabe y la guarda tendra que actuar.", flush=True)
        else:
            print("\n1) FASE 0 — topes mecanicos (empuje suave, para al no responder)...", flush=True)
            lo = find_limit(gb, -args.probe_u, args, samples, "ABAJO")
            hi = find_limit(gb, +args.probe_u, args, samples, "ARRIBA")
        travel = hi - lo
        center = (hi + lo) / 2.0
        print(">>> RECORRIDO: %.1f .. %.1f  = %.1f deg de rango  (centro %.1f)"
              % (lo, hi, travel, center), flush=True)
        if travel < 10:
            print("!! recorrido absurdo (%.1f deg). ¿El gimbal esta trabado? Abortado." % travel)
            return
        if args.probe:
            print("\n--probe: solo los topes. Fin.")
            return

        # ---------------- FASE A: escalones ----------------
        print("\n2) FASE A — escalones (curva estatica: velocidad vs comando)...", flush=True)
        levels = [int(x) for x in args.levels.split(",") if x.strip()]
        curve = {}
        goto(gb, center, args, samples)
        for L in levels:
            # +L y luego -L: SIMETRICO => vuelve solo al punto de partida, sin reposicionar.
            # (Antes esto barria de tope a tope y era inviable con un eje lento: a 4.5 deg/s
            #  cruzar 150 deg son 33 s, y a u=100 mas de 5 minutos.)
            r_up = step(gb, +L, lo, hi, args, samples)
            r_dn = step(gb, -L, lo, hi, args, samples)
            curve[L] = (r_up, r_dn)
            print("   u=%+4d -> %s  |  u=%+4d -> %s   (gpitch=%.1f)"
                  % (L, ("subiendo %.2f deg/s" % r_up) if r_up else "sin medida",
                     -L, ("bajando %.2f deg/s" % r_dn) if r_dn else "sin medida",
                     gb.gp if gb.gp is not None else 0), flush=True)
            if gb.gp is not None and abs(gb.gp - center) > 20:
                goto(gb, center, args, samples)      # recentra si derivo

        # ---------------- FASE B: PRBS dimensionado con los datos de arriba ----------------
        usable = [(L, (abs(u) + abs(d)) / 2) for L, (u, d) in curve.items() if u and d]
        if not usable:
            print("!! la FASE A no midio ninguna velocidad. Sin PRBS.");
        elif args.prbs_bits > 0:
            print("\n3) FASE B — PRBS, amplitud calculada del recorrido REAL...", flush=True)
            room = travel / 2.0 - args.margin                  # espacio a cada lado del centro
            best = None
            for L, rate in sorted(usable):
                exc = rate * args.prbs_bit_secs * PRBS_EXC     # excursion acotada de la m-seq
                fits = exc <= room * args.exc_frac
                print("   u=%+4d: rate %.1f deg/s -> excursion %.1f deg  (cabe en %.1f? %s)"
                      % (L, rate, exc, room * args.exc_frac, "SI" if fits else "no"), flush=True)
                if fits:
                    best = (L, rate, exc)
            if best is None:
                print("   !! ni el nivel mas bajo cabe. Baja --prbs-bit-secs o sube --levels.")
            else:
                amp, rate, exc = best
                print(">>> PRBS: amplitud %+d (rate %.1f deg/s), excursion prevista %.1f deg"
                      " sobre %.1f de sitio" % (amp, rate, exc, room), flush=True)
                gc = goto(gb, center, args, samples)
                gb.hold(1.0, 0)
                # NO arrancar sin estar centrado: desde otro sitio la excursion calculada
                # ya no cabe y el PRBS acaba contra el tope (fue exactamente el bug del
                # goto con doble pump).
                if gc is None or abs(gc - center) > 5.0:
                    print("   !! no se pudo centrar (%.1f vs %.1f): PRBS ABORTADO para no"
                          " estampar el gimbal." % (gc or 0, center), flush=True)
                else:
                    bits = prbs7(args.prbs_bits)
                    n_guard = 0
                    for i, b in enumerate(bits):
                        u = amp * b
                        t_end = time.time() + args.prbs_bit_secs
                        while time.time() < t_end:
                            # GUARDA DURA: cerca de un tope, no empujar MAS hacia el. Se
                            # modifica el comando; NO se hace un pump extra (seria la misma
                            # carrera de antes).
                            g = gb.gp
                            u_eff = u
                            if g is not None and ((u > 0 and g > hi - args.margin)
                                                  or (u < 0 and g < lo + args.margin)):
                                u_eff = 0; n_guard += 1
                            gnew, t = gb.pump(u_eff)      # <- el UNICO pump de la vuelta
                            if gnew is not None:
                                samples.append(dict(t_s=round(t - args.t0, 4), u_cmd=int(u_eff),
                                                    phase="PRBS", gpitch=gnew))
                        if i % 20 == 0:
                            print("   PRBS %3d/%d  gpitch=%.1f" % (i, len(bits), gb.gp or 0),
                                  flush=True)
                    if n_guard:
                        print("   (la guarda de topes actuo en %d muestras)" % n_guard, flush=True)
        print("\n4) centrando la camara...", flush=True)
        goto(gb, center, args, samples)
        gb.hold(0.5, 0)
    except KeyboardInterrupt:
        print("\n!! Ctrl+C -> parando y centrando", flush=True)
        try:
            gb.hold(0.5, 0)
        except Exception:
            pass
    finally:
        if s.sock:
            s.sock.close()

    if samples:
        with open(args.out, "w", newline="", encoding="utf-8") as fh:
            fh.write("# fecha: %s\n" % time.strftime("%Y-%m-%d %H:%M:%S"))
            fh.write("# Ts_nominal_s: 0.1 (Push Position 0x04/0x05 ~9.9Hz)\n")
            fh.write("# signo: u_cmd + = ARRIBA (gpitch sube); - = ABAJO\n")
            fh.write("# comando_enviado: 1024 + u_cmd (0x04/0x01, rango 363..1685)\n")
            w = csv.DictWriter(fh, fieldnames=["t_s", "u_cmd", "phase", "gpitch"])
            w.writeheader(); w.writerows(samples)
        print(">>> CSV: %d muestras -> %s" % (len(samples), args.out))
        print(">>> Exporta:  .\\.venv\\Scripts\\python tools\\neo_control\\sysid_gimbal.py"
              " --export %s" % args.out)


def export_xlsx(csv_path, xlsx_path, ts=TS):
    """CSV -> Excel. Hoja 'datos' = 3 columnas (tiempo, entrada, salida) en rejilla uniforme.
    El gimbal NO envuelve, asi que la salida va tal cual (sin des-envolver)."""
    import numpy as np
    from openpyxl import Workbook
    meta, body = {}, []
    for ln in open(csv_path, encoding="utf-8"):
        if ln.startswith("#"):
            k, _, v = ln[1:].partition(":"); meta[k.strip()] = v.strip()
        else:
            body.append(ln)
    rows = list(csv.DictReader(body))
    if not rows:
        print("!! el CSV no tiene muestras"); return
    t = np.array([float(r["t_s"]) for r in rows])
    u = np.array([float(r["u_cmd"]) for r in rows])
    g = np.array([float(r["gpitch"]) for r in rows])
    ph = [r["phase"] for r in rows]

    wb = Workbook()
    grid = np.arange(t[0], t[-1], ts)
    g_g = np.interp(grid, t, g)
    idx = np.clip(np.searchsorted(t, grid, side="right") - 1, 0, len(u) - 1)
    u_g = u[idx]; ph_g = [ph[i] for i in idx]

    # --- hoja 'datos': 3 columnas, entrada CRUDA (unidades de stick) ---
    ws = wb.active; ws.title = "datos"
    ws.append(["tiempo_s", "entrada_u", "salida_gpitch_deg"])
    for i in range(len(grid)):
        ws.append([round(float(grid[i]), 3), int(u_g[i]), round(float(g_g[i]), 3)])
    for c, w_ in (("A", 12), ("B", 12), ("C", 18)):
        ws.column_dimensions[c].width = w_

    # --- hoja 'datos_lin': 3 columnas, entrada LINEALIZADA (deg/s) ---
    # RECOMENDADA para ajustar. La estatica del gimbal es zona muerta 50 + expo 1.704
    # (R2=0.999998), y su ganancia local varia x5.8 entre u=100 y u=660. Un ajuste lineal
    # sobre la entrada CRUDA promedia esa no-linealidad y solo vale cerca de la amplitud
    # a la que se ajusto (aqui, el PRBS fue a |u|=500). Con la entrada ya linealizada la
    # planta es limpia y el PID vale en TODO el rango -- aplicando la inversion en el
    # controlador. (En el yaw esto era opcional: alli la ganancia solo variaba x1.9.)
    ws = wb.create_sheet("datos_lin")
    r_lin = np.sign(u_g) * EXPO_A_FIT * np.maximum(np.abs(u_g) - DEADZONE_FIT, 0.0) ** EXPO_N_FIT
    ws.append(["tiempo_s", "entrada_rate_deg_s", "salida_gpitch_deg"])
    for i in range(len(grid)):
        ws.append([round(float(grid[i]), 3), round(float(r_lin[i]), 4), round(float(g_g[i]), 3)])
    for c, w_ in (("A", 12), ("B", 20), ("C", 18)):
        ws.column_dimensions[c].width = w_

    ws = wb.create_sheet("info")
    sw = [r for r in rows if r["phase"].startswith("SWEEP")]
    info = [
        ("QUE ES ESTO", "Identificacion del GIMBAL (tilt de camara) del DJI Neo. TODO EN TIERRA."),
        ("", ""),
        ("  tiempo_s", "paso FIJO %.3f s. IMPUESTO por el Push Position 0x04/0x05 (9.90 Hz)" % ts),
        ("  entrada_u", "deflexion sobre el centro. El comando enviado es 1024 + entrada_u"),
        ("", "(0x04/0x01, rango fisico 363..1685). SIGNO: + = ARRIBA, - = ABAJO."),
        ("  salida_gpitch_deg", "angulo de la camara. 0 = al frente, negativo = mirando abajo."),
        ("", "Resolucion 0.1 deg (int16/10). NO envuelve: es un eje con TOPES."),
        ("", ""),
        ("MODELO ESPERADO", "cmd --> rate --[1/s]--> angulo  = INTEGRADOR (igual que el yaw)."),
        ("", "0x04/0x01 es un comando de VELOCIDAD: 1024 = quieta."),
        ("", "Al ser integrador, un P/PD ya da error nulo a escalon: la I SOBRA."),
        ("", "En el yaw la I costo 12.5% de overshoot y un ciclo limite (EXP-036)."),
        ("", ""),
        ("OJO — TOPES", "a diferencia del yaw, este eje tiene limite mecanico. Las fases"),
        ("", "LIMIT_* son la busqueda de topes; ignorarlas al identificar."),
        ("FASE SWEEP*", "barridos a velocidad constante -> curva estatica (rate vs comando)."),
        ("FASE PRBS", "excitacion rica -> dinamica/retardo. Amplitud dimensionada para que"),
        ("", "la excursion (rate*Tb*7, acotada en la m-secuencia) quepa en el recorrido."),
        ("FASE GOTO", "reposicionamiento entre fases. NO identificar con esto (es lazo cerrado)."),
        ("", ""),
    ]
    for k, v in meta.items():
        info.append(("meta: " + k, v))
    for r in info:
        ws.append(list(r))
    ws.column_dimensions["A"].width = 20; ws.column_dimensions["B"].width = 92

    ws = wb.create_sheet("raw")
    ws.append(["t_s", "u_cmd", "phase", "gpitch", "dt_s"])
    for i, r in enumerate(rows):
        ws.append([float(r["t_s"]), int(float(r["u_cmd"])), r["phase"], float(r["gpitch"]),
                   round(t[i] - t[i - 1], 4) if i else 0.0])
    wb.save(xlsx_path)
    print(">>> Excel: %d crudas, %d en rejilla de %.0f ms -> %s"
          % (len(rows), len(grid), ts * 1000, xlsx_path))


def main():
    ap = argparse.ArgumentParser(description="Identificacion del gimbal del Neo (EN TIERRA)")
    ap.add_argument("--probe", action="store_true", help="SOLO la FASE 0 (topes). Correr esto primero")
    ap.add_argument("--lo", type=float, default=None,
                    help="tope INFERIOR conocido (deg), salta la FASE 0. El usuario midio -90 con "
                         "la app. Requiere --hi. OJO: lo que permite la app puede NO ser lo que "
                         "alcanza nuestro path -> --probe lo confirma")
    ap.add_argument("--hi", type=float, default=None,
                    help="tope SUPERIOR conocido (deg), salta la FASE 0. El usuario midio +60")
    ap.add_argument("--export", metavar="CSV", default=None, help="CSV -> .xlsx (usa el .venv)")
    ap.add_argument("--out", default="sysid_gimbal.csv")
    # FASE 0
    ap.add_argument("--probe-u", dest="probe_u", type=float, default=250.0,
                    help="velocidad para buscar los topes. SUAVE a proposito")
    ap.add_argument("--still-deg", dest="still_deg", type=float, default=0.15,
                    help="movimiento por debajo del cual se considera 'quieto' (tope)")
    ap.add_argument("--still-secs", dest="still_secs", type=float, default=0.7,
                    help="segundos quieto para declarar tope")
    ap.add_argument("--limit-timeout", dest="limit_timeout", type=float, default=45.0,
                    help="tope de tiempo para hallar CADA limite. 45s, no 12: a --probe-u 250 "
                         "el gimbal va a ~4 deg/s y cruzar 90 deg tarda 22s. Con 12s el "
                         "--probe se quedaba a medias y fingia un recorrido de 46 deg")
    ap.add_argument("--margin", type=float, default=6.0,
                    help="grados de respeto a los topes: NUNCA se empuja dentro de este margen")
    # FASE A
    ap.add_argument("--levels", default="100,200,300,400,500,660",
                    help="niveles |u| de los barridos (se hacen + y -)")
    ap.add_argument("--step-secs", dest="step_secs", type=float, default=8.0,
                    help="duracion de cada escalon. Debe bastar para que la velocidad se "
                         "establezca Y para medirla con 0.1 deg de resolucion si el eje es lento")
    # FASE B
    ap.add_argument("--prbs-bits", dest="prbs_bits", type=int, default=127, help="0 = sin PRBS")
    ap.add_argument("--prbs-bit-secs", dest="prbs_bit_secs", type=float, default=0.3,
                    help="periodo de bit. Con Ts=0.1 son 3 muestras/bit: no bajar de 0.2")
    ap.add_argument("--exc-frac", dest="exc_frac", type=float, default=0.7,
                    help="fraccion del sitio disponible que puede ocupar la excursion del PRBS")
    args = ap.parse_args()
    if args.export:
        export_xlsx(args.export, os.path.splitext(args.export)[0] + ".xlsx"); return
    run(args)


if __name__ == "__main__":
    main()
