r"""
sysid_yaw.py — IDENTIFICACION DEL EJE DE YAW del DJI Neo (para diseñar un PID).

QUE MIDE
  Excita el stick de yaw (ch3) con una señal conocida y registra el ANGULO de yaw que
  devuelve el OSD del dron. Salida: un CSV crudo (durante el vuelo) que se convierte a
  Excel (.xlsx) con --export. Con eso se ajusta el modelo y se diseña el PID.

LA PLANTA (importante para interpretar los datos)
  El FC del Neo trata el stick de yaw como COMANDO DE VELOCIDAD, no de par: con stick
  constante el dron gira a velocidad ~constante (medido: defl 250 -> ~6 deg/s;
  defl 600 -> ~18 deg/s, EXP-032/033). Es decir:

        stick --[ zona muerta ]--> yaw_rate --[ 1/s ]--> yaw_angle

  => G(s) ~ K / (s (tau s + 1)): un INTEGRADOR con dinamica de primer orden dentro.
  CONSECUENCIA 1: identificar sobre la VELOCIDAD (planta estable), no sobre el angulo
  (que sobre un integrador deriva sin acotar). El 1/s se añade analiticamente despues.
  CONSECUENCIA 2: al ser un integrador, un P o PD ya da error nulo ante escalon; la
  accion integral normalmente NO hace falta y suele empeorar el margen de fase.

  ZONA MUERTA: ajustando rate = k (|u| - d) a los dos puntos conocidos sale d ~ 75,
  k ~ 0.034 deg/s por unidad. Ademas el yaw-hold del FC frena por debajo de ~220 (EXP-033).
  PERO la zona muerta NO afecta al PRBS de este ensayo: es BANG-BANG a saturacion
  (u = -660 / +660), o sea que |u| = 660 SIEMPRE y la zona muerta se atraviesa de golpe,
  sin reposar nunca en ella. Solo mordería a un PRBS de amplitud baja. Ventajas del
  bang-bang: maxima relacion señal/ruido (~6 deg por bit contra 0.1 deg de cuantizacion)
  y, al ser simetrico, SIN deriva neta (el desbalance de una m-secuencia de 127 bits es
  1 bit ~ 6 deg). A cambio identifica tau a UNA sola amplitud (saturacion) -> la curva de
  ganancia y la zona muerta las da la FASE A (escalones), no el PRBS.

MUESTREO (medido sobre 'Novena captura.pcap', 677 muestras / 68.6 s)
  El OSD 0x03/0x43 llega a 9.86 Hz => Ts = 100 ms. NO se puede muestrear mas rapido:
  es la tasa a la que el dron reporta. yaw viene int16/10 => resolucion 0.1 deg.
  Hay huecos ocasionales (max visto 1.1 s) -> se registra el timestamp REAL de cada
  muestra y --export remuestrea a rejilla uniforme aparte.

SECUENCIA DEL VUELO
  despegue -> ascenso (fuera del efecto suelo) -> FASE A (escalera de escalones: da la
  curva estatica y la zona muerta reales) -> FASE B (PRBS sesgado: da la dinamica tau)
  -> aterrizaje. El yaw NO traslada => girar es seguro en interior.

USO
  .\neo.ps1 sysid_yaw.py                        DRY: valida la secuencia en TIERRA (sin
                                                despegar), sticks NEUTRO. Escribe CSV vacio.
  .\neo.ps1 sysid_yaw.py --fly --armed-ok       VUELO REAL. Escribe sysid_yaw.csv.
  .\.venv\Scripts\python tools\neo_control\sysid_yaw.py --export sysid_yaw.csv
                                                CSV -> sysid_yaw.xlsx (3 hojas). Necesita
                                                el .venv (numpy + openpyxl).

SIGNO: ch3+ => giro a la DERECHA => el yaw del OSD SUBE (EXP-028).
Ctrl+C = ATERRIZA (throttle-min).
"""
import argparse, csv, os, subprocess, sys, time, socket

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import neo_udp as N
import flight as F

NEUTRAL = F.NEUTRAL

# Rango REAL del stick (flight.py:92): 364 .. 1684, centro 1024 => deflexion +-660.
# Saturar el yaw = ch3 364 (maximo IZQUIERDA) / ch3 1684 (maximo DERECHA).
MAX_DEFL = 660


# ---------------------------------------------------------------- señal de excitacion
def prbs7(n):
    """m-secuencia de 7 bits (x^7 + x^6 + 1), periodo 127. Devuelve +-1.
    Es determinista: la misma secuencia sale en cada vuelo -> vuelos comparables."""
    reg = 0x7f
    seq = []
    for _ in range(n):
        seq.append(1 if (reg & 1) else -1)
        fb = ((reg >> 0) ^ (reg >> 1)) & 1
        reg = (reg >> 1) | (fb << 6)
    return seq


def build_segments(args):
    """Plan de excitacion como lista de (etiqueta, u, segundos). u = deflexion del stick
    de yaw sobre 1024 (signo: + derecha). El plan es simetrico -> el dron acaba mirando
    ~al mismo sitio."""
    segs = [("SETTLE", 0, args.settle_air)]

    # --- FASE A: escalera. Da la NO LINEALIDAD ESTATICA (zona muerta + ganancia).
    # Sin neutro entre escalones: la rotacion acumulada da igual (el yaw no traslada) y
    # asi cada escalon parte de una velocidad ya establecida = mas datos utiles/segundo.
    levels = [int(x) for x in args.step_levels.split(",") if x.strip()]
    for u in levels:
        segs.append(("STEP+%d" % u, +u, args.step_secs))
    segs.append(("ZERO", 0, 2.0))
    for u in levels:
        segs.append(("STEP-%d" % u, -u, args.step_secs))
    segs.append(("ZERO", 0, 2.0))

    # --- FASE B: PRBS. Da la DINAMICA (tau). Por defecto BANG-BANG a saturacion
    # (bias=0, amp=660): u alterna entre -660 y +660 = stick ch3 364 <-> 1684.
    if args.prbs_bits > 0:
        for b in prbs7(args.prbs_bits):
            segs.append(("PRBS", int(args.prbs_bias + b * args.prbs_amp), args.prbs_bit_secs))
        segs.append(("ZERO", 0, 2.0))
        # Un PRBS SESGADO (bias != 0) gira siempre hacia el mismo lado y acumula rotacion:
        # se deshace con el sesgo opuesto. El bang-bang (bias=0) es simetrico y no la
        # necesita (deriva neta ~1 bit).
        if args.prbs_bias != 0:
            segs.append(("UNWIND", -args.prbs_bias, args.prbs_bits * args.prbs_bit_secs * 0.5))
            segs.append(("ZERO", 0, 2.0))
    return segs


def u_at(segs, t):
    """(u, etiqueta) en el instante t del plan; (None, None) si el plan ya termino."""
    acc = 0.0
    for label, u, secs in segs:
        if t < acc + secs:
            return u, label
        acc += secs
    return None, None


def plan_secs(segs):
    return sum(s[2] for s in segs)


# ---------------------------------------------------------------- lazo de adquisicion
def sysid_loop(f, segs, real, alt_max, samples):
    """Streamea el stick de yaw segun el plan y registra CADA muestra del OSD.
    Mantiene las mismas pistas que _fly_loop (suscripcion/modo/autoridad/keepalive), que
    es lo que sostiene el enganche con el FC. Devuelve motivo de fin: 'plan'|'alt'|'land'.
    En DRY los sticks van NEUTRO (registra igual -> valida la cadena en tierra)."""
    t0 = time.time()
    total = plan_secs(segs)
    nt = {"stick": 0.0, "mode": 0.0, "auth": 0.0, "ka": 0.0, "sub": 0.0, "rep": 0.0}
    yaw_prev = None
    yaw_unwrap = 0.0
    n0 = len(samples)
    while True:
        now = time.time()
        t = now - t0
        u, label = u_at(segs, t)
        if u is None:
            return "plan"
        sticks = (1024, 1024, 1024, 1024 + int(u)) if real else NEUTRAL

        if f.serial and now >= nt["sub"]:
            f.sub13(); nt["sub"] = now + 0.2
        if now >= nt["mode"]:
            f.set_mode(); nt["mode"] = now + 0.1
        if now >= nt["stick"]:
            f.stick(*sticks); nt["stick"] = now + 0.05      # 20 Hz
        if now >= nt["auth"]:
            f.authority(); nt["auth"] = now + 1.0
        if now >= nt["ka"]:
            f.s.keepalive(); nt["ka"] = now + 0.5

        f.s.sock.settimeout(0.03)
        try:
            d, a = f.s.sock.recvfrom(65535)
        except (socket.timeout, BlockingIOError):
            d = None
        if not (d and a[0] == N.DRONE[0]):
            continue
        b = N.find_battery_dynamic(d)
        if b:
            f.last_batt = b
        o = N.find_osd_general(d)
        if not o:
            continue

        # --- una muestra del OSD: es NUESTRO instante de muestreo (Ts ~ 100 ms) ---
        yaw = o["yaw"]
        if yaw_prev is None:
            yaw_unwrap = yaw
        else:
            dy = yaw - yaw_prev                       # des-envolver el salto +-180
            if dy > 180: dy -= 360
            elif dy < -180: dy += 360
            yaw_unwrap += dy
        yaw_prev = yaw
        samples.append(dict(
            t_s=round(t, 4), u_cmd=int(u), phase=label,
            yaw_deg=yaw, yaw_unwrap_deg=round(yaw_unwrap, 3),
            height_m=o["height_m"], motor_on=int(o["motor_on"]),
            volt_mv=(f.last_batt or {}).get("voltage_mv", ""),
        ))

        if real and o["height_m"] > alt_max:
            print("!! altura %.1fm > tope %.1fm -> CORTO el plan" % (o["height_m"], alt_max), flush=True)
            return "alt"
        if real and not o["motor_on"] and t > 3.0:
            print("!! motores OFF inesperado -> CORTO el plan", flush=True)
            return "land"
        if now >= nt["rep"]:
            print("  t+%5.1f [%-8s] u=%+5d  yaw=%7.1f (acum %8.1f)  alt=%.1fm  n=%d"
                  % (t, label, u, yaw, yaw_unwrap, o["height_m"], len(samples) - n0), flush=True)
            nt["rep"] = now + 1.0


def climb(f, args, samples):
    """Sube fuera del efecto suelo antes de excitar. Misma logica validada en mapflight
    (EXP-033): el auto-despegue se queda ~0.7m y su altitude-hold resiste empujes timidos;
    hay que empujar FUERTE y SOSTENIDO, y solo tras estabilizar (alt>=0.3), nunca a ciegas.
    OJO: hay ~0.8m de inercia tras soltar el stick (cortar en 1.5 -> acaba en ~2.3)."""
    d_thr = max(0, min(int(args.climb_thr), 640))
    stick = (1024, 1024, 1024 + d_thr, 1024)
    t0 = time.time()
    push_start = None
    nt = {"stick": 0.0, "mode": 0.0, "auth": 0.0, "ka": 0.0, "sub": 0.0, "rep": 0.0}
    osd = None
    while True:
        now = time.time(); t = now - t0
        h = osd["height_m"] if osd else 0.0
        if h >= 0.3 and push_start is None:
            push_start = now
        pushed = (now - push_start) if push_start else 0.0
        if (push_start and pushed >= args.climb_secs) or h >= args.alt or t > args.climb_max:
            print(">>> ascenso fin: alt=%.1fm (empuje %.1fs)" % (h, pushed), flush=True)
            return osd
        if f.serial and now >= nt["sub"]:
            f.sub13(); nt["sub"] = now + 0.2
        if now >= nt["mode"]:
            f.set_mode(); nt["mode"] = now + 0.1
        if now >= nt["stick"]:
            f.stick(*(stick if h >= 0.3 else NEUTRAL)); nt["stick"] = now + 0.05
        if now >= nt["auth"]:
            f.authority(); nt["auth"] = now + 1.0
        if now >= nt["ka"]:
            f.s.keepalive(); nt["ka"] = now + 0.5
        f.s.sock.settimeout(0.03)
        try:
            d, a = f.s.sock.recvfrom(65535)
        except (socket.timeout, BlockingIOError):
            d = None
        if d and a[0] == N.DRONE[0]:
            o = N.find_osd_general(d)
            if o: osd = o
        if now >= nt["rep"]:
            print("  t+%5.1f [CLIMB] alt=%.1fm" % (t, h), flush=True); nt["rep"] = now + 1.0


# ---------------------------------------------------------------- salida
CSV_COLS = ["t_s", "u_cmd", "phase", "yaw_deg", "yaw_unwrap_deg", "height_m", "motor_on", "volt_mv"]


def write_csv(path, samples, meta):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        for k, v in meta.items():
            fh.write("# %s: %s\n" % (k, v))
        w = csv.DictWriter(fh, fieldnames=CSV_COLS)
        w.writeheader()
        for s in samples:
            w.writerow(s)
    print(">>> CSV: %d muestras -> %s" % (len(samples), path), flush=True)


def export_xlsx(csv_path, xlsx_path, ts=0.1):
    """CSV crudo -> Excel con 3 hojas: info / raw / resampled_100ms.
    Requiere el .venv (numpy + openpyxl)."""
    import numpy as np
    from openpyxl import Workbook

    meta, rows = {}, []
    with open(csv_path, encoding="utf-8") as fh:
        lines = [ln for ln in fh]
    body = []
    for ln in lines:
        if ln.startswith("#"):
            k, _, v = ln[1:].partition(":")
            meta[k.strip()] = v.strip()
        else:
            body.append(ln)
    rows = list(csv.DictReader(body))
    if not rows:
        print("!! el CSV no tiene muestras (¿fue un DRY?)"); return

    t = np.array([float(r["t_s"]) for r in rows])
    u = np.array([float(r["u_cmd"]) for r in rows])
    yw = np.array([float(r["yaw_unwrap_deg"]) for r in rows])
    ph = [r["phase"] for r in rows]

    wb = Workbook()

    # --- hoja 'datos': LA HOJA DE TRABAJO. 3 columnas, rejilla uniforme de Ts.
    # tiempo | entrada (deflexion del stick) | salida (yaw des-envuelto, grados).
    # La salida va DES-ENVUELTA a proposito: el yaw crudo del OSD salta +180 <-> -180
    # y esos saltos arruinarian cualquier ajuste.
    ws = wb.active; ws.title = "datos"
    grid = np.arange(t[0], t[-1], ts)
    yw_g = np.interp(grid, t, yw)                       # salida: interpolacion lineal
    idx = np.clip(np.searchsorted(t, grid, side="right") - 1, 0, len(u) - 1)
    u_g = u[idx]                                        # entrada: retencion de orden cero
    ph_g = [ph[i] for i in idx]                         # (es un comando por tramos)
    ws.append(["tiempo_s", "entrada_u", "salida_yaw_deg"])
    for i in range(len(grid)):
        ws.append([round(float(grid[i]), 3), int(u_g[i]), round(float(yw_g[i]), 3)])
    for c, wd in (("A", 12), ("B", 12), ("C", 16)):
        ws.column_dimensions[c].width = wd

    # --- hoja info: todo lo que hace falta para interpretar las columnas ---
    ws = wb.create_sheet("info")
    info = [
        ("QUE ES ESTO", "Identificacion del eje de YAW del DJI Neo (stick ch3 -> angulo de yaw)"),
        ("", ""),
        ("HOJA 'datos'", "3 columnas, rejilla uniforme. Es la hoja de trabajo."),
        ("  tiempo_s", "segundos desde el inicio de la excitacion. Paso FIJO Ts = %.3f s" % ts),
        ("  entrada_u", "deflexion del stick de yaw sobre el centro. UNIDADES CRUDAS del stick:"),
        ("", "el valor enviado al dron es 1024 + entrada_u. Rango fisico +-%d." % MAX_DEFL),
        ("", "SIGNO: + = DERECHA. Saturacion: -%d = ch3 %d (max izq), +%d = ch3 %d (max der)"
             % (MAX_DEFL, 1024 - MAX_DEFL, MAX_DEFL, 1024 + MAX_DEFL)),
        ("  salida_yaw_deg", "angulo de yaw en grados, DES-ENVUELTO (continuo, acumulado)."),
        ("", "El yaw crudo del OSD salta +180 <-> -180; aqui esos saltos YA estan quitados."),
        ("", "Resolucion del sensor: 0.1 deg (int16/10 del OSD)."),
        ("", ""),
        ("MODELO MEDIDO", "(EXP-034, ajustado sobre ESTOS datos. Reemplaza la hipotesis previa"),
        ("", " de 'zona muerta d~75', que era FALSA: venia de solo 2 puntos sucios.)"),
        ("", ""),
        ("  1) ESTATICA", "NO hay zona muerta: hay EXPO. rate = 49.5 * (|u|/660)^1.351  [deg/s]"),
        ("", "R2=0.9997 sobre los 6 escalones (el modelo lineal+zona muerta da R2=0.9930 y"),
        ("", "yerra el doble en u=100). Fondo de escala: ~49.5 deg/s con el stick saturado."),
        ("", "La ganancia local dr/du varia x1.9 entre u=100 y u=660 -> conviene INVERTIR el"),
        ("", "expo antes del PID:  u = 660 * sign(r) * (|r|/49.5)^0.740"),
        ("", ""),
        ("  2) DINAMICA", "yaw[k] = yaw[k-1] + Ts * r_sp[k-1]   =>   G(z) = Ts*z^-1 / (1 - z^-1)"),
        ("", "INTEGRADOR PURO + 1 muestra. Ese z^-1 es el ZOH normal de discretizar 1/s:"),
        ("", "NO hay retardo de transporte extra medible (WiFi+OSD quedan por debajo de Ts)."),
        ("", "tau del lazo interno de velocidad: NO OBSERVABLE (cambiarlo de 0.08 a 0 mueve el"),
        ("", "R2 en 0.00006). Se establece en <2 muestras => irrelevante a Ts=100ms."),
        ("", "Matiz: en conmutacion rapida (PRBS) la ganancia efectiva cae a ~0.84 del valor"),
        ("", "estatico -> hay algo de retraso sub-muestra sin modelar. Dejar margen."),
        ("", ""),
        ("  3) PARA EL PID", "Con G(z)=Ts*z^-1/(1-z^-1) y un P puro: z_lc = 1 - Ts*Kp"),
        ("", "=> estable para 0 < Kp < 20 ; deadbeat en Kp=10. Con la caida de ganancia a 0.84"),
        ("", "y el expo sin invertir, el margen real es menor: empezar bajo (Kp ~ 2-4)."),
        ("", "Al ser la planta un integrador, un P/PD ya da error nulo a escalon: la I sobra."),
        ("", ""),
        ("Ts", "%.3f s (%.1f Hz). IMPUESTO por el OSD del dron: no se puede muestrear mas" % (ts, 1 / ts)),
        ("", "rapido. Medido sobre 'Novena captura.pcap': 677 muestras / 68.6 s = 9.86 Hz."),
        ("", "Hay huecos ocasionales (max visto 1.1 s) -> ver dt_s en la hoja 'raw'."),
        ("", ""),
        ("hoja 'raw'", "muestras TAL CUAL llegaron, con su dt real (variable) y altura/bateria"),
        ("hoja 'detalle'", "misma rejilla + fase + yaw_rate_dps (derivada por diferencia central;"),
        ("", "el OSD NO trae velocidad angular: es DERIVADA, no medida. Ruido de"),
        ("", "cuantizacion ~ 0.1deg/0.1s = 1 deg/s -> filtrar si se identifica sobre el rate)."),
        ("", ""),
        ("FASE A", "escalera de escalones -> curva ESTATICA. Salio limpisima (R2=0.9997) y es"),
        ("", "la fuente FIABLE del expo. Con escalones de 3s el retardo NO se puede leer aqui."),
        ("FASE B", "PRBS bang-bang -> es la unica fase donde el retardo SI se identifica"),
        ("", "(d=1 muestra: R2=0.81 contra 0.53 con d=0). tau no: es sub-muestra."),
        ("", ""),
    ]
    for k, v in meta.items():
        info.append(("meta: " + k, v))
    for r in info:
        ws.append(list(r))
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 96

    # --- hoja raw ---
    ws = wb.create_sheet("raw")
    ws.append(CSV_COLS + ["dt_s"])
    for i, r in enumerate(rows):
        dt = (t[i] - t[i - 1]) if i else 0.0
        ws.append([float(r["t_s"]), int(float(r["u_cmd"])), r["phase"], float(r["yaw_deg"]),
                   float(r["yaw_unwrap_deg"]), float(r["height_m"]), int(r["motor_on"]),
                   r["volt_mv"], round(dt, 4)])

    # --- hoja detalle: la misma rejilla + la fase y la velocidad derivada (por si se
    # quiere identificar sobre el rate en vez de sobre el angulo). Secundaria.
    ws = wb.create_sheet("detalle")
    rate = np.gradient(yw_g, ts)                        # diferencia central
    ws.append(["tiempo_s", "entrada_u", "fase", "salida_yaw_deg", "yaw_rate_dps"])
    for i in range(len(grid)):
        ws.append([round(float(grid[i]), 3), int(u_g[i]), ph_g[i],
                   round(float(yw_g[i]), 3), round(float(rate[i]), 3)])

    wb.save(xlsx_path)
    print(">>> Excel: %d muestras crudas, %d en rejilla de %.0f ms -> %s"
          % (len(rows), len(grid), ts * 1000, xlsx_path))

    # resumen de la curva estatica: util para leer la zona muerta de un vistazo
    print("\n--- FASE A: velocidad estacionaria por escalon (ultimo 50% de cada tramo) ---")
    print("     u_cmd    rate_dps")
    seen = {}
    for i in range(len(grid)):
        if not ph_g[i].startswith("STEP"):
            continue
        seen.setdefault(ph_g[i], []).append((int(u_g[i]), float(rate[i])))
    for k in sorted(seen, key=lambda x: seen[x][0][0]):
        vals = seen[k]
        tail = [r for _, r in vals[len(vals) // 2:]]     # ignora el transitorio
        if tail:
            print("     %+5d    %+7.2f" % (vals[0][0], sum(tail) / len(tail)))


# ---------------------------------------------------------------- main
def run(args):
    real = args.fly and args.armed_ok
    segs = build_segments(args)
    print("=" * 68)
    print("  sysid_yaw.py — %s" % ("VUELO REAL (identificacion de yaw)" if real
                                   else "DRY (sticks NEUTRO, no despega)"))
    print("  FASE A escalones: %s  (%.1fs c/u)" % (args.step_levels, args.step_secs))
    lo_u, hi_u = args.prbs_bias - args.prbs_amp, args.prbs_bias + args.prbs_amp
    print("  FASE B PRBS: %d bits x %.2fs = %.1fs  |  u: %+d <-> %+d  (ch3: %d <-> %d)%s"
          % (args.prbs_bits, args.prbs_bit_secs, args.prbs_bits * args.prbs_bit_secs,
             lo_u, hi_u, 1024 + lo_u, 1024 + hi_u,
             "  BANG-BANG a saturacion" if args.prbs_bias == 0 and args.prbs_amp == MAX_DEFL else ""))
    print("  plan total: %.1fs de excitacion  |  Ts = 100ms (OSD del dron)" % plan_secs(segs))
    print("=" * 68)
    if args.fly and not args.armed_ok:
        print("!! --fly requiere --armed-ok. Abortado."); return
    if real:
        hi = [args.prbs_bias + args.prbs_amp, args.prbs_bias - args.prbs_amp]
        if max(abs(x) for x in hi) > MAX_DEFL:
            print("!! el PRBS pide |u|=%d > tope del stick (%d). Abortado."
                  % (max(abs(x) for x in hi), MAX_DEFL)); return
        # la zona muerta solo muerde si la señal REPOSA en amplitud baja: el bang-bang
        # (|u| constante = amp) la atraviesa sin pararse y no le afecta.
        if min(abs(x) for x in hi) < 220:
            print("!! AVISO: el PRBS reposa en |u|=%d, por debajo del umbral del yaw-hold"
                  " (~220). La FASE B puede salir sucia." % min(abs(x) for x in hi))
        print("\n" + "!" * 68)
        print(" VUELO REAL. El dron VA A GIRAR SOBRE SI MISMO casi todo el vuelo (el yaw no")
        print(" traslada, pero necesita espacio libre alrededor y supervision).")
        print(" Sube a ~%.1fm. Ctrl+C = ATERRIZAR." % args.alt)
        print("!" * 68 + "\n")

    samples = []
    reason = "?"
    # Type5Session (NO es context manager: se cierra a mano). Su send_command ya pacea a
    # la ventana RX del dron y retransmite -> por eso esta herramienta se apoya en flight.py
    # y no en el camino crudo de mapflight.py (ver EXP-032).
    s = N.Type5Session()
    if not s.open():
        print("!! sin ACK al hello -> revisa WiFi del Neo / DJI Fly cerrado."); return
    print("hello -> ACK. Sesion abierta (seed=0x%04x session=0x%04x)" % (s.seed, s.session))
    f = F.Flight(s, args.lat, args.lon)
    try:
        F.run_common(f, args)
        if real:
            print("3) DESPEGUE (AUTO_FLY)...", flush=True)
            f.auto_fly()
            climb(f, args, samples)
        else:
            print("3) DRY: sin despegue; corriendo el plan con sticks NEUTRO.", flush=True)
        print("4) EXCITACION (%.1fs)..." % plan_secs(segs), flush=True)
        reason = sysid_loop(f, segs, real, args.alt_max, samples)
        print(">>> plan terminado (%s)" % reason, flush=True)
    except KeyboardInterrupt:
        print("\n!! Ctrl+C -> ATERRIZANDO", flush=True)
        reason = "ctrl-c"
    finally:
        if real:
            print("5) ATERRIZAJE (throttle-min)...", flush=True)
            try:
                ok = f.descend(args.land)
                print(">>> touchdown %s" % ("CONFIRMADO" if ok else "NO confirmado (revisar)"),
                      flush=True)
            except KeyboardInterrupt:
                print("!! Ctrl+C durante el aterrizaje. Si sigue en el aire: boton del dron.",
                      flush=True)
        if s.sock:
            s.sock.close()

    meta = {
        "fecha": time.strftime("%Y-%m-%d %H:%M:%S"),
        "modo": "VUELO" if real else "DRY",
        "fin": reason,
        "Ts_nominal_s": "0.1 (OSD 0x03/0x43 ~9.9Hz; dt real en la hoja raw)",
        "signo": "u_cmd + = DERECHA = yaw sube",
        "stick_enviado": "1024 + u_cmd (canal ch3)",
        "fase_A_niveles": args.step_levels,
        "fase_A_secs": args.step_secs,
        "prbs": "7-bit m-seq, %d bits, bit=%.2fs, bias=%d, amp=%d"
                % (args.prbs_bits, args.prbs_bit_secs, args.prbs_bias, args.prbs_amp),
    }
    write_csv(args.out, samples, meta)
    if samples:
        print(">>> Exporta a Excel:  .\\.venv\\Scripts\\python tools\\neo_control\\sysid_yaw.py"
              " --export %s" % args.out)


def main():
    ap = argparse.ArgumentParser(description="Identificacion del eje de yaw del DJI Neo")
    ap.add_argument("--fly", action="store_true", help="VUELO REAL (motores)")
    ap.add_argument("--armed-ok", dest="armed_ok", action="store_true",
                    help="2do candado: confirma area despejada + supervisado")
    ap.add_argument("--export", metavar="CSV", default=None,
                    help="convierte un CSV de vuelo a .xlsx (usa el .venv: numpy+openpyxl)")
    ap.add_argument("--out", default="sysid_yaw.csv", help="CSV de salida del vuelo")
    ap.add_argument("--lat", type=float, default=None)
    ap.add_argument("--lon", type=float, default=None)
    ap.add_argument("--settle", type=float, default=4.0, help="settle en tierra antes de despegar")
    ap.add_argument("--settle-air", dest="settle_air", type=float, default=3.0,
                    help="hover neutro tras el ascenso, antes de excitar")
    ap.add_argument("--land", type=float, default=14.0, help="ventana de aterrizaje")
    # --- ascenso (validado en EXP-033) ---
    ap.add_argument("--alt", type=float, default=1.5, help="corte del ascenso (m). OJO: ~0.8m de inercia despues")
    ap.add_argument("--alt-max", dest="alt_max", type=float, default=3.5,
                    help="guarda dura: si se supera, se corta el plan y se aterriza")
    ap.add_argument("--climb-thr", dest="climb_thr", type=float, default=450.0)
    ap.add_argument("--climb-secs", dest="climb_secs", type=float, default=8.0)
    ap.add_argument("--climb-max", dest="climb_max", type=float, default=16.0)
    # --- FASE A: escalones ---
    ap.add_argument("--step-levels", dest="step_levels", default="100,200,300,400,500,660",
                    help="niveles |u| de la escalera (se hacen + y -). Dan la curva estatica. "
                         "660 = saturacion (el mismo punto donde opera el PRBS bang-bang)")
    ap.add_argument("--step-secs", dest="step_secs", type=float, default=3.0,
                    help="duracion de cada escalon (debe bastar para que la velocidad se establezca)")
    # --- FASE B: PRBS ---
    ap.add_argument("--prbs-bits", dest="prbs_bits", type=int, default=127,
                    help="bits de la m-secuencia (127 = periodo completo de 7 bits). 0 = sin PRBS")
    ap.add_argument("--prbs-bit-secs", dest="prbs_bit_secs", type=float, default=0.3,
                    help="periodo de bit (s). Con Ts=0.1 son 3 muestras/bit: no bajar de 0.2")
    ap.add_argument("--prbs-amp", dest="prbs_amp", type=int, default=MAX_DEFL,
                    help="amplitud del PRBS (u = bias +- amp). Default %d = SATURACION del stick "
                         "(bang-bang: ch3 salta entre %d y %d)" % (MAX_DEFL, 1024 - MAX_DEFL, 1024 + MAX_DEFL))
    ap.add_argument("--prbs-bias", dest="prbs_bias", type=int, default=0,
                    help="punto de operacion del PRBS. Default 0 = bang-bang simetrico (sin deriva "
                         "neta). Solo sesgar si se quiere identificar a amplitud baja, donde la "
                         "zona muerta SI molesta")
    args = ap.parse_args()

    if args.export:
        xlsx = os.path.splitext(args.export)[0] + ".xlsx"
        export_xlsx(args.export, xlsx)
        return
    run(args)


if __name__ == "__main__":
    main()
